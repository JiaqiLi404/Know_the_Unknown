# @Time : 2023/12/21 16:42
# @Author : Li Jiaqi
# @Description :
import datasets
import openpyxl
import os
import copy
import random
from typing import List
import jsonlines
import json


#################### Loading HF datasets ####################
def load_hagrid_dataset():
    hagrid = datasets.load_dataset("miracl/hagrid", split="train")
    return hagrid


#################### Loading xls datasets ####################
def load_xls_dataset(filename, has_title=True):
    sum_res = dict()
    workbook = openpyxl.load_workbook(filename)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        res = list(worksheet.rows)
        if has_title:
            res = res[1:]
        for i in range(len(res)):
            res[i] = list(res[i])
            res[i] = [x.value for x in res[i]]
            res[i] = [x for x in res[i] if x is not None]
        res = [x for x in res if len(x) > 0]
        sum_res[sheet] = res
    return sum_res, workbook


def formulate_dataset(dataset):
    for k, v in dataset.items():
        dataset[k] = [x[0] for x in v]


def save_xls_dataset(dataset, filename, title=['query', 'context', 'ori_answer', "answer1", "answer2"]):
    workbook = openpyxl.Workbook()
    for k, v in dataset.items():
        workbook.create_sheet(k)
        worksheet = workbook[k]
        for i in range(len(title)):
            worksheet.cell(row=1, column=i + 1).value = title[i]
        for i in range(len(dataset[k])):
            for j in range(len(dataset[k][i])):
                worksheet.cell(row=i + 2, column=j + 1).value = str(dataset[k][i][j])
    workbook.save(filename)


#################### Loading json datasets ####################
def load_jsonl_dataset(filename):
    with open(filename) as f:
        for item in jsonlines.Reader(f):
            yield item
def lead_nq_dataset():
    dataset = load_jsonl_dataset(os.path.join("qa_datasets", "json_analyzer_files", "simplified-nq-train.jsonl"))
    for d in dataset:
        query, context, answer = d["question_text"], d["document_text"], d['annotations'][0]['short_answers']
        answer_text = "Not Provided"
        if len(answer) > 0:
            answer_text = " ".join(context.split(" ")[answer[0]['start_token']:answer[0]['end_token']])
        yield query, context, answer_text

def load_json_dataset(filename):
    with open(filename, 'r') as f:
        dataset = json.loads(f.read())
    return dataset

#################### Transform datasets ####################
def clean_context(context):
    lines = context.split("\\n")
    for i in range(len(lines) - 1):
        if lines[i].endswith((".", "?", "!")) or lines[i].istitle() or lines[i].isupper():
            lines[i] += "\n"
        else:
            lines[i] += " "
    context = "".join(lines)

    # lines = context.split("\n")
    # pattern = re.compile(r"^[^\n]*-?\d+(\.\d+)?(/\d+)?[^.,;:?!]*$")
    # new_lines = [line for line in lines if not pattern.match(line)]
    # context = "\n".join(new_lines)

    spliter = "\n\n" if context.count("\n\n") > context.count("\n \n") else "\n \n"
    context = context.split(spliter)
    context = [x for x in context if len(x) > 0]
    while len(context) > 5:
        lengths = [len(i) for i in context]
        min_index = lengths.index(min(lengths))
        if min_index == len(context) - 1:
            context[-2] = context[-2] + "\n" + context[-1]
            context.pop()
        else:
            context[min_index] = context[min_index] + "\n" + context[min_index + 1]
            context.pop(min_index + 1)
    context = [f"[{str(i + 1)}] " + context[i] for i in range(len(context))]

    context = "\n\n\n".join(context)
    return context


def multiple_split(text, alphas=['\n', ' '], split_reference=False, split_min_length=1, keep_alphas=False):
    text_list = []
    while text[1] in alphas:
        text = text[0] + text[2:]
        if len(text) == 0:
            return []
    pointer = 0
    for i in range(len(text) - 1):
        if pointer < i and text[i] in alphas and len(text[pointer:i].strip()) > 0:
            # To avoid floats
            if '.' in alphas and text[i] == '.' and 0 < i < len(text) - 1 and str.isdigit(text[i - 1]) and str.isdigit(
                    text[i + 1]):
                continue
            # To avoid abbreviations, such as U.S.
            if len(text[pointer:i].strip()) < split_min_length or (
                    len(text_list) > 0 and len(text_list[-1]) < split_min_length):
                if pointer > 0:
                    pointer = pointer - 1
                if len(text_list) > 0:
                    text_list[-1] += text[pointer:i].strip() if not keep_alphas else text[pointer:i + 1].strip()
                else:
                    text_list.append(text[pointer:i].strip() if not keep_alphas else text[pointer:i + 1].strip())
            else:
                text_list.append(text[pointer:i].strip() if not keep_alphas else text[pointer:i + 1].strip())
            pointer = i + 1
        if split_reference and text[i:i + 3].startswith('[') and text[i:i + 3].endswith(']') and len(
                text[pointer:i].strip()) > 0:
            text_list.append(text[pointer:i].strip())
            pointer = i
    text_list.append(text[pointer:].strip())
    return text_list


def traceback_contexts(citations: List, context: str):
    def __pair_context(answer_sen, context_sen, threshold=0.73):
        answer_pointer = 0
        context_pointer = 0
        while answer_pointer < len(answer_sen) and context_pointer < len(context_sen):
            if answer_sen[answer_pointer].lower() == context_sen[context_pointer].lower():
                answer_pointer += 1
                context_pointer += 1
            elif answer_pointer + 1 < len(answer_sen) and answer_sen[answer_pointer + 1].lower() == context_sen[
                context_pointer].lower():
                answer_pointer += 2
                context_pointer += 1
            else:
                context_pointer += 1
        if answer_pointer >= len(answer_sen) * threshold:
            return True
        return False

    pured_context = multiple_split(context)
    pured_context = " ".join(pured_context)
    context_sentences = multiple_split(pured_context, alphas=['.', '?', '!'], split_reference=True, split_min_length=10)

    for citation_dict in citations:
        citation_list = citation_dict['citation']
        neg_sources = copy.deepcopy(context_sentences)
        if len(citation_list) == 0:
            neg_source_num = min(len(neg_sources), 5)
            citation_dict['neg_source'] = random.sample(neg_sources, neg_source_num)
            continue
        for citation in citation_list:
            # search the paired context sentence for each citation
            for context_sentence in context_sentences:
                if __pair_context(citation, context_sentence):
                    citation_dict['source'].append(context_sentence)
                    break
        for source in citation_dict['source']:
            if source in neg_sources:
                neg_sources.remove(source)
        if len(neg_sources) > 0:
            neg_source_num = min(len(neg_sources), 5)
            citation_dict['neg_source'] = random.sample(neg_sources, neg_source_num)
        if len(citation_dict['citation']) != len(citation_dict['source']):
            print('XXXXXXXX BUG : Some citations are not found in the context XXXXXXXX')
            print('citation:')
            print(citation_dict['citation'])
            print('source:')
            print(citation_dict['source'])
            print('context:')
            print(context_sentences)
            print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX\n\n\n')

    return citations


def separate_citations(answer_with_citation):
    answer_with_citation_list = []
    citation_flag = False
    pointer = 0
    for i in range(len(answer_with_citation) - 1):
        if not citation_flag and answer_with_citation[i:i + 2] == ']<':
            answer_with_citation_list.append(answer_with_citation[pointer:i + 1])
            citation_flag = True
            pointer = i + 1
        if citation_flag and answer_with_citation[i:i + 1] == '>':
            answer_with_citation_list.append(answer_with_citation[pointer:i + 1])
            citation_flag = False
            pointer = i + 1
    answer_with_citation_list.append(answer_with_citation[pointer:])

    answer_with_citation_list_temp = []
    for j in range(len(answer_with_citation_list)):
        if answer_with_citation_list[j].startswith('<') and answer_with_citation_list[j].endswith('>'):
            answer_with_citation_list_temp.append(answer_with_citation_list[j])
        else:
            answer_with_citation_list_temp.extend(
                multiple_split(answer_with_citation_list[j], alphas=['.', '?', '!'], split_min_length=10,
                               keep_alphas=True))
    citations = []
    for j in range(len(answer_with_citation_list_temp)):
        if j + 1 < len(answer_with_citation_list_temp) and answer_with_citation_list_temp[j + 1].startswith('<') and \
                answer_with_citation_list_temp[j + 1].endswith('>'):
            if answer_with_citation_list_temp[j + 1].endswith('...>'):
                answer_with_citation_list_temp[j + 1] = answer_with_citation_list_temp[j + 1][:-4] + '>'
            citations.append({
                "answer": answer_with_citation_list_temp[j],
                "citation": multiple_split(answer_with_citation_list_temp[j + 1][1:-1], alphas=['.', '?', '!', ';'],
                                           split_min_length=10),
                "source": [],
                "neg_source": []
            })
        elif not answer_with_citation_list_temp[j].startswith('<') or not answer_with_citation_list_temp[j].endswith(
                '>'):
            citations.append({
                "answer": answer_with_citation_list_temp[j],
                "citation": [],
                "source": [],
                "neg_source": []
            })
        # if '..' in citations[-1]['citation']:
        #     print('answer:')
        #     print(answer_with_citation_list_temp[j])
        #     print('citation:')
        #     print(answer_with_citation_list_temp[j+1])
        #     print('----------------------------------\n\n\n')
    return citations
